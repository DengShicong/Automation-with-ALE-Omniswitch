#!/usr/bin/env python3
# -*- coding:utf-8 -*-
"""
ALE网络运维工具包 - 设备巡检模块
支持ALE设备的tech-support命令和FTP文件传输
"""

import os
import sys
import ftplib
import time
from datetime import datetime
from netmiko import ConnectHandler
from openpyxl.reader.excel import load_workbook
from multiprocessing.pool import ThreadPool

# 导入配置
try:
    from ale_config import get_device_credentials, get_download_config
    CONFIG_AVAILABLE = True
except ImportError:
    CONFIG_AVAILABLE = False
    print("警告: ale_config.py不可用，使用默认配置")


class ALEInspection:
    """ALE网络运维工具包 - 设备巡检类"""
    
    def __init__(self):
        self.device_file = "template.xlsx"  # 使用现有的xlsx文件
        self.pool = ThreadPool(10)
        self.success = []
        self.fail = []
        self.log_dir = "LOG"
        self.logtime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        
        # 创建LOG目录
        if not os.path.exists(self.log_dir):
            os.makedirs(self.log_dir)
            print(f"创建LOG目录: {self.log_dir}")
    
    def load_excel(self):
        """加载Excel文件"""
        try:
            wb = load_workbook(self.device_file)
            return wb
        except FileNotFoundError:
            print(f"Excel文件不存在: {self.device_file}")
            return None
    
    def get_device_info(self):
        """获取设备信息"""
        try:
            wb = self.load_excel()
            if not wb:
                return

            ws1 = wb[wb.sheetnames[0]]
            for row in ws1.iter_rows(min_row=2, max_col=9):
                if str(row[1].value).strip() == '#':
                    continue

                device_type = str(row[8].value) if row[8].value else ""

                # 获取命令列表，如果工作表不存在则返回空列表
                cmd_list = []
                if device_type:
                    try:
                        sheet_name = device_type.lower().strip()
                        if sheet_name in wb.sheetnames:
                            cmd_list = self.get_cmd_info(wb[sheet_name])
                        else:
                            print(f"警告: 设备类型 '{device_type}' 对应的工作表不存在，将使用空命令列表")
                    except Exception as e:
                        print(f"获取设备 {row[2].value} 的命令列表失败: {e}")

                info_dict = {
                    'ip': row[2].value,
                    'protocol': row[3].value,
                    'port': row[4].value,
                    'username': row[5].value,
                    'password': row[6].value,
                    'secret': row[7].value,
                    'device_type': device_type,
                    'cmd_list': cmd_list
                }
                yield info_dict

        except Exception as e:
            print(f"读取设备信息错误: {e}")
    
    def get_cmd_info(self, cmd_sheet):
        """获取命令信息"""
        cmd_list = []
        try:
            if cmd_sheet is None:
                return []

            for row in cmd_sheet.iter_rows(min_row=2, max_col=2):
                if row[0].value and str(row[0].value).strip() != "#" and row[1].value:
                    cmd = str(row[1].value).strip()
                    if cmd:  # 确保命令不为空
                        cmd_list.append(cmd)

            print(f"从工作表 '{cmd_sheet.title}' 读取到 {len(cmd_list)} 个命令")
            return cmd_list

        except Exception as e:
            print(f"读取命令信息错误: {e}")
            return []

    def get_vendor_name(self, device_type):
        """根据设备类型获取厂商名称"""
        device_type = device_type.lower()

        vendor_mapping = {
            'cisco_ios': 'Cisco',
            'cisco_xe': 'Cisco',
            'cisco_xr': 'Cisco',
            'cisco_nxos': 'Cisco',
            'huawei': '华为',
            'huawei_vrp': '华为',
            'hp_comware': 'H3C',
            'h3c_comware': 'H3C',
            'ruijie_os': '锐捷',
            'juniper': 'Juniper',
            'juniper_junos': 'Juniper',
            'arista_eos': 'Arista',
            'fortinet': 'Fortinet',
            'paloalto_panos': 'Palo Alto',
            'dell_force10': 'Dell',
            'extreme': 'Extreme',
            'alcatel_aos': 'ALE',
            'alcatel_sros': 'Nokia'
        }

        return vendor_mapping.get(device_type, device_type.upper())
    
    def connect_device(self, host):
        """连接设备"""
        try:
            if host['protocol'].lower().strip() == 'ssh':
                host['port'] = host['port'] if (host['port'] not in [22, None]) else 22
            elif host['protocol'].lower().strip() == 'telnet':
                host['port'] = host['port'] if (host['port'] not in [23, None]) else 23
                host['device_type'] = host['device_type'] + '_telnet'
            
            # 移除不需要的参数
            connect_params = {
                'device_type': host['device_type'],
                'host': host['ip'],
                'username': host['username'],
                'password': host['password'],
                'port': host['port']
            }
            
            if host['secret']:
                connect_params['secret'] = host['secret']
            
            if 'huawei' in host['device_type']:
                connect_params['conn_timeout'] = 15
            
            connect = ConnectHandler(**connect_params)
            return connect
            
        except Exception as e:
            print(f"连接设备失败 {host['ip']}: {e}")
            self.fail.append(host['ip'])
            return None
    
    def execute_ale_tech_support(self, connection, device_ip, username=None, password=None):
        """执行ALE设备的tech-support命令并下载日志文件"""
        try:
            print(f"开始执行ALE设备 {device_ip} 的tech-support命令...")

            # 执行tech-support命令
            output = connection.send_command("show tech-support", delay_factor=2)
            print(f"tech-support命令执行完成: {device_ip}")

            # 获取配置信息
            if CONFIG_AVAILABLE:
                config = get_download_config()
                wait_time = config['files']['wait_time_after_tech_support']
                log_files = config['files']['tech_support_files']
                # 获取设备认证信息
                download_username, download_password = get_device_credentials(device_ip, username, password)
            else:
                wait_time = 10
                log_files = ["tech_support_layer3.log", "tech_support_layer2.log", "tech_support.log"]
                download_username, download_password = username, password

            # 等待文件生成
            print(f"等待日志文件生成: {wait_time}秒")
            time.sleep(wait_time)

            print(f"使用认证信息下载文件: 用户={download_username}")

            # 下载日志文件
            downloaded_files = []
            for log_file in log_files:
                if self.download_file_via_tftp(device_ip, log_file, download_username, download_password):
                    downloaded_files.append(log_file)
            
            if downloaded_files:
                print(f"成功下载 {device_ip} 的日志文件: {downloaded_files}")
                return True
            else:
                print(f"未能下载 {device_ip} 的任何日志文件")
                return False
                
        except Exception as e:
            print(f"执行tech-support失败 {device_ip}: {e}")
            return False
    
    def download_file_via_tftp(self, device_ip, filename, username=None, password=None):
        """尝试多种方式下载文件"""
        print(f"开始下载文件: {device_ip}:{filename}")

        # 方式1: 尝试FTP下载（首选）
        print("尝试方式1: FTP下载")
        if self.download_file_via_ftp(device_ip, filename, username, password):
            return True

        # 方式2: 尝试TFTP下载（如果有TFTP客户端）
        print("尝试方式2: TFTP下载")
        try:
            from tftp_downloader import TFTPClient
            tftp_client = TFTPClient(device_ip)

            device_log_dir = os.path.join(self.log_dir, f"{device_ip}_{self.logtime}")
            if not os.path.exists(device_log_dir):
                os.makedirs(device_log_dir)

            local_filename = f"{device_ip}_{filename}"
            local_path = os.path.join(device_log_dir, local_filename)

            if tftp_client.download_file(filename, local_path):
                print(f"✓ TFTP下载成功: {local_path}")
                return True

        except Exception as e:
            print(f"✗ TFTP下载失败: {e}")

        # 所有方式都失败，创建备用记录
        print(f"所有下载方式都失败，创建备用记录: {filename}")
        self.create_backup_record(device_ip, filename, "所有下载方式(FTP/TFTP)都失败")
        return False
    
    def download_file_via_netmiko(self, connection, device_ip, filename):
        """通过netmiko连接下载文件（使用设备命令）"""
        try:
            print(f"尝试通过设备命令获取文件内容: {device_ip}:{filename}")

            # 创建设备专用目录
            device_log_dir = os.path.join(self.log_dir, f"{device_ip}_{self.logtime}")
            if not os.path.exists(device_log_dir):
                os.makedirs(device_log_dir)

            # 本地文件名包含设备IP
            local_filename = f"{device_ip}_{filename}"
            local_path = os.path.join(device_log_dir, local_filename)

            # 尝试使用ALE设备的文件查看命令
            file_commands = [
                f"more {filename}",
                f"cat {filename}",
                f"show file {filename}",
                f"file show {filename}"
            ]

            file_content = None
            successful_command = None

            for cmd in file_commands:
                try:
                    print(f"  尝试命令: {cmd}")
                    output = connection.send_command(cmd, delay_factor=2)

                    # 检查输出是否包含有效内容
                    if output and len(output) > 50 and "No such file" not in output and "not found" not in output.lower():
                        file_content = output
                        successful_command = cmd
                        break

                except Exception as e:
                    print(f"  命令失败: {cmd} - {e}")
                    continue

            if file_content:
                # 保存文件内容
                with open(local_path, 'w', encoding='utf-8') as f:
                    f.write(f"# ALE设备文件内容\n")
                    f.write(f"# 设备IP: {device_ip}\n")
                    f.write(f"# 文件名: {filename}\n")
                    f.write(f"# 获取命令: {successful_command}\n")
                    f.write(f"# 获取时间: {datetime.now()}\n")
                    f.write("# " + "=" * 50 + "\n\n")
                    f.write(file_content)

                print(f"✓ 文件内容获取成功: {local_path}")
                return True
            else:
                print(f"✗ 无法获取文件内容: {filename}")
                return False

        except Exception as e:
            print(f"✗ 文件获取失败 {device_ip}/{filename}: {e}")
            return False



    def download_file_via_ftp(self, device_ip, filename, username=None, password=None):
        """通过FTP下载文件（首选方案）"""
        try:
            # 使用传入的认证信息
            ftp_server = device_ip
            ftp_user = username if username else "admin"
            ftp_password = password if password else "password"

            print(f"尝试FTP连接: {ftp_server} (用户: {ftp_user})")

            # 创建设备专用目录
            device_log_dir = os.path.join(self.log_dir, f"{device_ip}_{self.logtime}")
            if not os.path.exists(device_log_dir):
                os.makedirs(device_log_dir)

            # 本地文件名包含设备IP
            local_filename = f"{device_ip}_{filename}"
            local_path = os.path.join(device_log_dir, local_filename)

            # 尝试FTP下载
            with ftplib.FTP(ftp_server) as ftp:
                ftp.login(ftp_user, ftp_password)
                with open(local_path, 'wb') as local_file:
                    ftp.retrbinary(f'RETR {filename}', local_file.write)

            print(f"✓ FTP下载成功: {local_path}")
            return True

        except Exception as e:
            print(f"✗ FTP下载失败 {device_ip}/{filename}: {e}")
            return False

    def create_backup_record(self, device_ip, filename, error_msg):
        """创建备用记录文件"""
        try:
            device_log_dir = os.path.join(self.log_dir, f"{device_ip}_{self.logtime}")
            if not os.path.exists(device_log_dir):
                os.makedirs(device_log_dir)

            backup_file = os.path.join(device_log_dir, f"{device_ip}_{filename}.failed")
            with open(backup_file, 'w', encoding='utf-8') as f:
                f.write(f"ALE设备日志文件下载失败记录\n")
                f.write(f"设备IP: {device_ip}\n")
                f.write(f"文件名: {filename}\n")
                f.write(f"时间: {datetime.now()}\n")
                f.write(f"错误信息: {error_msg}\n")
                f.write("=" * 50 + "\n")
                f.write("可能的解决方案:\n")
                f.write("1. 检查设备FTP服务是否启用\n")
                f.write("2. 确认用户名密码正确\n")
                f.write("3. 检查文件是否存在于设备根目录\n")
                f.write("4. 手动从设备下载文件\n")
                f.write("5. 检查网络连接和防火墙设置\n")
                f.write("6. 确认FTP端口21是否开放\n")

            print(f"创建备用记录: {backup_file}")
            return True

        except Exception as e:
            print(f"创建备用记录失败: {e}")
            return False
    
    def execute_regular_commands(self, connection, device_ip, cmd_list, device_type):
        """执行常规命令并保存回显到单个文件"""
        try:
            device_log_dir = os.path.join(self.log_dir, f"{device_ip}_{self.logtime}")
            if not os.path.exists(device_log_dir):
                os.makedirs(device_log_dir)

            print(f"开始执行 {len(cmd_list)} 个命令: {device_ip}")
            successful_commands = 0
            failed_commands = 0

            # 创建统一的命令输出文件
            output_filename = f"{device_ip}_{device_type}_commands_output.txt"
            output_file_path = os.path.join(device_log_dir, output_filename)

            with open(output_file_path, 'w', encoding='utf-8') as output_file:
                # 写入文件头
                output_file.write(f"设备: {device_ip}\n")
                output_file.write(f"设备类型: {device_type}\n")
                output_file.write(f"执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                output_file.write(f"命令总数: {len(cmd_list)}\n")
                output_file.write("=" * 80 + "\n\n")

                for i, cmd in enumerate(cmd_list, 1):
                    try:
                        print(f"  [{i}/{len(cmd_list)}] 执行命令: {cmd}")
                        command_output = connection.send_command(cmd, delay_factor=2)

                        # 写入命令和输出到统一文件
                        output_file.write(f"[命令 {i}] {cmd}\n")
                        output_file.write("-" * 60 + "\n")
                        output_file.write(command_output)
                        output_file.write("\n" + "=" * 80 + "\n\n")

                        print(f"  ✓ 命令完成: {cmd}")
                        successful_commands += 1

                    except Exception as e:
                        print(f"  ✗ 命令失败: {cmd} - {e}")
                        failed_commands += 1

                        # 写入错误信息到统一文件
                        output_file.write(f"[命令 {i}] {cmd} - 执行失败\n")
                        output_file.write("-" * 60 + "\n")
                        output_file.write(f"错误信息: {str(e)}\n")
                        output_file.write("=" * 80 + "\n\n")

                # 写入文件尾
                output_file.write(f"\n执行汇总:\n")
                output_file.write(f"成功命令: {successful_commands}\n")
                output_file.write(f"失败命令: {failed_commands}\n")
                output_file.write(f"完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

            print(f"命令执行汇总 {device_ip}: 成功 {successful_commands}, 失败 {failed_commands}")
            print(f"所有命令输出已保存到: {output_filename}")
            return successful_commands > 0

        except Exception as e:
            print(f"执行命令列表失败 {device_ip}: {e}")
            return False
    
    def inspect_device(self, host):
        """巡检单个设备"""
        device_ip = host['ip']
        device_type = host['device_type'].lower()
        
        print(f"开始巡检设备: {device_ip} ({device_type})")
        
        # 连接设备
        connection = self.connect_device(host)
        if not connection:
            return
        
        try:
            # 判断设备类型并执行相应操作
            if 'alcatel' in device_type or 'ale' in device_type:
                # ALE设备：只执行tech-support和下载日志
                print(f"检测到ALE设备: {device_ip} - 执行tech-support流程")

                tech_support_success = self.execute_ale_tech_support(
                    connection, device_ip, host['username'], host['password']
                )

                if tech_support_success:
                    print(f"✓ ALE设备 {device_ip} tech-support流程完成")
                else:
                    print(f"✗ ALE设备 {device_ip} tech-support流程失败")

            else:
                # 其他厂商设备：执行对应命令列表
                print(f"检测到{device_type}设备: {device_ip} - 执行命令列表")

                if host['cmd_list']:
                    regular_success = self.execute_regular_commands(connection, device_ip, host['cmd_list'], device_type)
                    if regular_success:
                        print(f"✓ 设备 {device_ip} 命令执行完成")
                    else:
                        print(f"✗ 设备 {device_ip} 命令执行失败")
                else:
                    print(f"! 设备 {device_ip} 没有配置命令列表")

            # 记录成功
            self.success.append(device_ip)
            print(f"设备处理完成: {device_ip}")

        except Exception as e:
            print(f"设备处理失败: {device_ip} - {e}")
            self.fail.append(device_ip)
            
        finally:
            connection.disconnect()
    
    def run_inspection(self):
        """运行完整运维流程"""
        print("=" * 60)
        print("开始网络设备运维")
        print(f"时间: {self.logtime}")
        print("=" * 60)

        start_time = datetime.now()

        # 获取设备列表并执行运维
        devices = list(self.get_device_info())
        if not devices:
            print("没有找到设备配置")
            return

        # 统计设备类型
        ale_devices = []
        other_devices = []

        for device in devices:
            device_type = device['device_type'].lower()
            if 'alcatel' in device_type or 'ale' in device_type:
                ale_devices.append(device)
            else:
                other_devices.append(device)

        # 统计各厂商设备数量
        vendor_count = {}
        for device in other_devices:
            vendor = self.get_vendor_name(device['device_type'])
            vendor_count[vendor] = vendor_count.get(vendor, 0) + 1

        print(f"发现 {len(devices)} 个设备:")
        print(f"  - ALE设备: {len(ale_devices)} 个 (执行tech-support流程)")
        for vendor, count in vendor_count.items():
            print(f"  - {vendor}设备: {count} 个 (执行命令列表)")
        print()

        # 并发执行运维
        for host in devices:
            self.pool.apply_async(self.inspect_device, args=(host,))

        self.pool.close()
        self.pool.join()
        
        end_time = datetime.now()
        
        # 打印结果
        print("\n" + "=" * 60)
        print("运维任务完成")
        print("=" * 60)
        print(f"总设备数: {len(devices)}")
        print(f"  - ALE设备: {len(ale_devices)} 个")
        print(f"  - 其他设备: {len(other_devices)} 个")
        print(f"成功设备: {len(self.success)}")
        print(f"失败设备: {len(self.fail)}")
        print(f"耗时: {(end_time - start_time).total_seconds():.2f}秒")

        if self.success:
            print("\n✓ 成功设备:")
            for device in self.success:
                # 判断设备类型
                device_info = next((d for d in devices if d['ip'] == device), None)
                if device_info:
                    device_type = device_info['device_type'].lower()
                    if 'alcatel' in device_type or 'ale' in device_type:
                        print(f"  ✓ {device} (ALE - tech-support)")
                    else:
                        vendor_name = self.get_vendor_name(device_type)
                        print(f"  ✓ {device} ({vendor_name} - 命令列表)")
                else:
                    print(f"  ✓ {device}")

        if self.fail:
            print("\n✗ 失败设备:")
            for device in self.fail:
                print(f"  ✗ {device}")

        # 压缩LOG文件夹
        self.compress_and_email(devices)
    
    def compress_and_email(self, devices):
        """为每个设备单独压缩并发送邮件"""
        try:
            import zipfile

            print(f"\n开始为每个设备创建压缩包...")
            zip_files = []

            # 为每个成功的设备创建单独的压缩包
            for device_ip in self.success:
                device_dir = os.path.join(self.log_dir, f"{device_ip}_{self.logtime}")

                if os.path.exists(device_dir):
                    # 创建设备专用压缩包
                    zip_filename = os.path.join(self.log_dir, f"{device_ip}_{self.logtime}.zip")

                    try:
                        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                            # 遍历设备目录中的所有文件
                            for root, _, files in os.walk(device_dir):
                                for file in files:
                                    file_path = os.path.join(root, file)
                                    # 在压缩包中保持相对路径
                                    arcname = os.path.relpath(file_path, self.log_dir)
                                    zipf.write(file_path, arcname)

                        zip_files.append(zip_filename)
                        file_size = os.path.getsize(zip_filename) / (1024 * 1024)
                        print(f"✓ {device_ip} 压缩完成: {os.path.basename(zip_filename)} ({file_size:.2f}MB)")

                    except Exception as e:
                        print(f"✗ {device_ip} 压缩失败: {e}")
                else:
                    print(f"! {device_ip} 目录不存在，跳过压缩")

            # 创建总体汇总压缩包（可选）
            if zip_files:
                summary_zip = os.path.join(self.log_dir, f"all_devices_{self.logtime}.zip")
                try:
                    with zipfile.ZipFile(summary_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        # 添加所有设备的压缩包到汇总包中
                        for zip_file in zip_files:
                            zipf.write(zip_file, os.path.basename(zip_file))

                        # 如果有失败设备，创建失败设备列表文件
                        if self.fail:
                            failed_list_path = os.path.join(self.log_dir, "failed_devices.txt")
                            with open(failed_list_path, 'w', encoding='utf-8') as f:
                                f.write(f"运维失败设备列表\n")
                                f.write(f"时间: {self.logtime}\n")
                                f.write("=" * 40 + "\n")
                                for device in self.fail:
                                    f.write(f"{device}\n")
                            zipf.write(failed_list_path, "failed_devices.txt")

                    zip_files.append(summary_zip)
                    summary_size = os.path.getsize(summary_zip) / (1024 * 1024)
                    print(f"✓ 汇总压缩包创建完成: {os.path.basename(summary_zip)} ({summary_size:.2f}MB)")

                except Exception as e:
                    print(f"✗ 汇总压缩包创建失败: {e}")

            if zip_files:
                print(f"\n总共创建了 {len(zip_files)} 个压缩包")

                # 发送邮件
                self.send_email_with_attachments(devices, zip_files)
            else:
                print("✗ 没有创建任何压缩包")

        except Exception as e:
            print(f"压缩和邮件发送过程出错: {e}")

    def send_email_with_attachments(self, devices, zip_files):
        """发送包含多个附件的邮件"""
        try:
            from send_email import send_email
            print("准备发送邮件...")

            # 准备设备信息用于邮件
            success_devices_info = []
            for device_ip in self.success:
                device_info = next((d for d in devices if d['ip'] == device_ip), None)
                if device_info:
                    device_type = device_info['device_type'].lower()
                    if 'alcatel' in device_type or 'ale' in device_type:
                        success_devices_info.append(f"{device_ip} (ALE)")
                    else:
                        vendor_name = self.get_vendor_name(device_type)
                        success_devices_info.append(f"{device_ip} ({vendor_name})")
                else:
                    success_devices_info.append(device_ip)

            # 检查附件大小
            total_size = sum(os.path.getsize(f) for f in zip_files if os.path.exists(f)) / (1024 * 1024)
            print(f"附件总大小: {total_size:.2f}MB")

            # 如果附件太大，只发送汇总包
            if total_size > 25:  # 25MB限制
                print("附件过大，只发送汇总压缩包")
                summary_files = [f for f in zip_files if 'all_devices_' in f]
                if summary_files:
                    zip_files = summary_files
                else:
                    # 如果没有汇总包，发送最小的几个文件
                    zip_files = sorted(zip_files, key=lambda x: os.path.getsize(x))[:3]
                    print(f"发送最小的 {len(zip_files)} 个压缩包")

            # 发送邮件，包含所有压缩包
            success = send_email(
                attachment_files=zip_files,
                success_devices=success_devices_info,
                failed_devices=self.fail,
                total_time=f"{len(self.success + self.fail)} 个设备"
            )

            if success:
                print("✓ 邮件发送成功!")
            else:
                print("✗ 邮件发送失败")

        except Exception as e:
            print(f"邮件发送失败: {e}")


def main():
    """主函数"""
    print("ALE网络运维工具包")
    print("支持tech-support命令和日志文件下载")
    print("=" * 60)
    
    # 检查Excel文件
    if not os.path.exists("template.xlsx"):
        print("错误: template.xlsx文件不存在")
        print("请确保Excel配置文件存在")
        return
    
    # 创建巡检实例并运行
    inspector = ALEInspection()
    inspector.run_inspection()


if __name__ == '__main__':
    main()
