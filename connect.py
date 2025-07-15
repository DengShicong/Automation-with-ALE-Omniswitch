from datetime import datetime
from multiprocessing import pool
from multiprocessing.pool import ThreadPool

import pandas as pd
from netmiko import ConnectHandler, NetmikoTimeoutException
from openpyxl.reader.excel import load_workbook
from paramiko.ssh_exception import AuthenticationException, SSHException


# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

class BackupConfig(object):


    def __init__(self):
        self.device_file = "template.xlsx"
        self.pool = ThreadPool(10)
        self.success = []
        self.fail = []
    def load_excel(self):
        try:
            wb=load_workbook(self.device_file)
            return wb
        except FileNotFoundError:
            print("{}File Not Found".format(self.device_file))

    def get_device_info(self):

         try:
             wb = self.load_excel()
             ws1 = wb[wb.sheetnames[0]]
             for row in ws1.iter_rows(min_row=2,max_col=9):
                 if str(row[1].value).strip() == '#':
                     continue
                 info_dict = {'ip': row[2].value,
                              'protocol': row[3].value,
                              'port': row[4].value,
                              'username': row[5].value,
                              'password': row[6].value,
                              'secret': row[7].value,
                              'device_type': row[8].value,
                              'cmd_list': self.get_cmd_info(wb[row[8].value.lower().strip()])

                 }
                 yield info_dict

         except Exception as e:
             print("ERROR:",e)

         finally:
             pass


    def get_cmd_info(self,cmd_sheet):
        cmd_list=[]
        try:
            for row in cmd_sheet.iter_rows(min_row=2,max_col=2):
                if str(row[0].value).strip() != '#' and row[1].value:
                    cmd_list.append(row[1].value.strip())

            return cmd_list

        except Exception as e:
            print("get_cmd_info Error:",e)


    def connectHandler(self,host):
        try:
            connect = ''
            if host['protocol'].lower().strip() == 'ssh':
                host['port'] = host['port'] if (host['port'] not in [22,None]) else 22
                host.pop('protocol'),host.pop('cmd_list')

                if 'huawei' in host['device_type']:
                    connect = ConnectHandler(**host,conn_timeout=15)
                else:
                    connect = ConnectHandler(**host)

            elif host['protocol'].lower().strip() == 'telnet':
                host['port'] = host['port'] if (host['port'] not in [23, None]) else 23
                host.pop('protocol'), host.pop('cmd_list')
                host['device_type'] = host['device_type'] + '_telnet'

                connect = ConnectHandler(**host,fast_cli=False)

            else:
                res = "{}_Not_Support_Protocol".format(host['ip'],host['protocol'])
                raise ValueError(res)

            return connect

        except NetmikoTimeoutException as e:
            res = "Failed connect: {}".format(host['ip'])
            print(res)
            self.fail.append(host['ip'])
            return None

        except AuthenticationException as e:
            res = "Failed Auth: {}".format(host['ip'])
            print(res)
            self.fail.append(host['ip'])
            return None

        except SSHException as e:
            res = "Failed SSH: {}".format(host['ip'])
            print(res)
            self.fail.append(host['ip'])
            return None

        except Exception as e:
            print("connectionHandler Failed: {} - {}".format(host['ip'], e))
            self.fail.append(host['ip'])
            return None

    def run_cmd(self,host,cmds,enable=False):
        enable = True if host['secret'] else False

        try:
            conn = self.connectHandler(host)

            if conn:
                hostname = conn.find_prompt()
                print(f"成功连接到设备: {host['ip']} ({hostname})")

                if cmds:
                    output = ''
                    for cmd in cmds:
                        try:
                            if enable:
                                conn.enable()
                                result = conn.send_command(cmd,strip_command=False,strip_prompt=False)
                                output += f"\n命令: {cmd}\n{result}\n"
                            else:
                                result = conn.send_command(cmd, strip_command=False, strip_prompt=False)
                                output += f"\n命令: {cmd}\n{result}\n"
                        except Exception as cmd_e:
                            print(f"命令执行失败 {cmd}: {cmd_e}")
                            output += f"\n命令: {cmd}\n错误: {cmd_e}\n"

                    print(f"设备 {host['ip']} 命令执行完成")
                    if host['ip'] not in self.success:
                        self.success.append(host['ip'])
                else:
                    print(f"设备 {host['ip']} 无命令需要执行")
                    if host['ip'] not in self.success:
                        self.success.append(host['ip'])

                conn.disconnect()

        except Exception as e:
            print(f"run_cmd Failed: {host['ip']} - {e}")
            if host['ip'] not in self.fail:
                self.fail.append(host['ip'])

    def connect_t(self):
        """连接测试方法"""
        start_time = datetime.now()

        hosts = self.get_device_info()
        for host in hosts:
            self.test_connection(host)

        end_time = datetime.now()
        print("连接测试完成,耗时:{:0.2f}s".format((end_time-start_time).total_seconds()))

    def test_connection(self, host):
        """测试单个设备连接"""
        try:
            conn = self.connectHandler(host)
            if conn:
                hostname = conn.find_prompt()
                print(f"连接测试成功: {host['ip']} - {hostname}")
                if host['ip'] not in self.success:
                    self.success.append(host['ip'])
                conn.disconnect()
            else:
                print(f"连接测试失败: {host['ip']}")
                if host['ip'] not in self.fail:
                    self.fail.append(host['ip'])
        except Exception as e:
            print(f"连接测试异常: {host['ip']} - {e}")
            if host['ip'] not in self.fail:
                self.fail.append(host['ip'])

    def connect_test(self):
        pass

    def connect(self):

        start_time = datetime.now()

        hosts = self.get_device_info()
        for host in hosts:
            #self.run_cmd(host,host['cmd_list'])
            self.pool.apply_async(self.run_cmd, args=(host,host['cmd_list']))
        self.pool.close()
        self.pool.join()

        end_time = datetime.now()
        print("complete,time:{:0.2f}s".format((end_time-start_time).total_seconds()))

if __name__ == '__main__':
    BackupConfig().connect()


